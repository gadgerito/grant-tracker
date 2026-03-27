"""
pages/import_deliverables.py — Download template, fill in Excel, upload to tracker with Gantt chart
"""

import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import json
import os
from datetime import date, datetime
from io import BytesIO

if not st.session_state.get("authenticated"):
    st.warning("🔒 Please sign in from the main page first.")
    st.stop()

def get_db():
    try:
        from pymongo import MongoClient
        uri = os.environ.get("MONGODB_URI", "") or st.secrets.get("MONGODB_URI", "")
        client = MongoClient(uri, serverSelectionTimeoutMS=5000)
        client.admin.command("ping")
        return client["grant_tracker"]
    except Exception as e:
        st.error(f"DB error: {e}")
        return None

def load_deliverables(db):
    try:
        docs = list(db["deliverables"].find({}, {"_id": 0}))
        return pd.DataFrame(docs) if docs else pd.DataFrame(columns=COLUMNS)
    except:
        return pd.DataFrame(columns=COLUMNS)

COLUMNS = [
    "id", "deliverable", "description", "assignee",
    "due_date", "status", "budget_allocated", "budget_spent",
    "milestone", "notes"
]

STATUS_OPTIONS = ["Not Started", "In Progress", "Under Review", "Complete", "Blocked"]

TEMPLATE_DATA = pd.DataFrame([
    {
        "deliverable": "Example: Submit Aim 1 Report",
        "description": "Write and submit the quarterly report for Aim 1",
        "assignee": "Jane Smith",
        "due_date": "2026-06-30",
        "status": "Not Started",
        "budget_allocated": 5000,
        "budget_spent": 0,
        "milestone": "Aim 1",
        "notes": "Coordinate with PI before submission"
    },
    {
        "deliverable": "Example: Stakeholder Meeting",
        "description": "Host quarterly stakeholder update meeting",
        "assignee": "John Doe",
        "due_date": "2026-07-15",
        "status": "Not Started",
        "budget_allocated": 1000,
        "budget_spent": 0,
        "milestone": "Aim 2",
        "notes": ""
    }
])

def create_template_excel():
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        TEMPLATE_DATA.to_excel(writer, index=False, sheet_name="Deliverables")
        wb = writer.book
        ws = writer.sheets["Deliverables"]

        # Style the header row
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Auto-width columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # Add instructions sheet
        instructions = pd.DataFrame({
            "Field": ["deliverable", "description", "assignee", "due_date", "status",
                      "budget_allocated", "budget_spent", "milestone", "notes"],
            "Required": ["Yes", "No", "No", "Yes", "Yes", "No", "No", "No", "No"],
            "Format": ["Text", "Text", "Text", "YYYY-MM-DD", 
                       "Not Started / In Progress / Under Review / Complete / Blocked",
                       "Number (no $ sign)", "Number (no $ sign)", "Text", "Text"],
            "Example": ["Submit Q1 Report", "Write and submit quarterly report", 
                        "Jane Smith", "2026-06-30", "Not Started",
                        "5000", "0", "Aim 1", "Coordinate with PI"]
        })
        instructions.to_excel(writer, index=False, sheet_name="Instructions")
        ws2 = writer.sheets["Instructions"]
        for cell in ws2[1]:
            cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        for col in ws2.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buffer.seek(0)
    return buffer

def create_export_excel(df):
    buffer = BytesIO()
    export_cols = [c for c in COLUMNS if c in df.columns and c != "id"]
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df[export_cols].to_excel(writer, index=False, sheet_name="Deliverables")
        wb = writer.book
        ws = writer.sheets["Deliverables"]
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    buffer.seek(0)
    return buffer

# ── UI ────────────────────────────────────────────────────────────────────────
st.title("📊 Import & Export Deliverables")
st.caption("Download the Excel template, fill it in, and upload it to add or update deliverables in bulk.")

db = get_db()
if db is None:
    st.stop()

existing_df = load_deliverables(db)

tab1, tab2, tab3 = st.tabs(["⬇️ Download Template", "⬆️ Upload & Import", "📤 Export Current"])

# ═══════════════════════════════════════════════════════════════════════
# TAB 1 — DOWNLOAD TEMPLATE
# ═══════════════════════════════════════════════════════════════════════
with tab1:
    st.subheader("⬇️ Download Excel Template")
    st.markdown("""
**How to use:**
1. Click **Download Template** below
2. Open the file in Excel or Google Sheets
3. Delete the example rows and add your deliverables
4. Save as `.xlsx`
5. Come back and upload it in the **Upload & Import** tab
""")

    st.info("📋 The template includes an **Instructions** sheet explaining each column.")

    try:
        template_buffer = create_template_excel()
        st.download_button(
            "⬇️ Download Excel Template",
            data=template_buffer,
            file_name="grant_deliverables_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )
    except ImportError:
        st.error("openpyxl not installed. Run: pip install openpyxl")
        st.code("deliverable, description, assignee, due_date, status, budget_allocated, budget_spent, milestone, notes")

    st.markdown("---")
    st.markdown("**Column reference:**")
    ref = pd.DataFrame({
        "Column": ["deliverable ✅", "description", "assignee", "due_date ✅", "status ✅",
                   "budget_allocated", "budget_spent", "milestone", "notes"],
        "Required": ["Yes", "No", "No", "Yes", "Yes", "No", "No", "No", "No"],
        "Example": ["Submit Q1 Report", "Write quarterly report", "Jane Smith",
                    "2026-06-30", "Not Started", "5000", "0", "Aim 1", "Coordinate with PI"]
    })
    st.dataframe(ref, use_container_width=True, hide_index=True)

# ═══════════════════════════════════════════════════════════════════════
# TAB 2 — UPLOAD & IMPORT
# ═══════════════════════════════════════════════════════════════════════
with tab2:
    st.subheader("⬆️ Upload Excel File")

    import_mode = st.radio(
        "Import mode",
        ["➕ Add new deliverables (keep existing)", "🔄 Replace ALL deliverables (full overwrite)"],
        horizontal=False
    )

    if "Replace" in import_mode:
        st.warning("⚠️ **Replace mode** will delete ALL existing deliverables and replace with the uploaded file. Make sure you have a backup!")

    uploaded = st.file_uploader("Upload your filled Excel file (.xlsx)", type=["xlsx"])

    if uploaded:
        try:
            upload_df = pd.read_excel(uploaded, sheet_name="Deliverables")
            st.success(f"✅ File loaded: **{len(upload_df)} rows found**")

            # Validate required columns
            required = ["deliverable", "due_date", "status"]
            missing = [c for c in required if c not in upload_df.columns]
            if missing:
                st.error(f"Missing required columns: {missing}")
                st.stop()

            # Validate status values
            invalid_status = upload_df[~upload_df["status"].isin(STATUS_OPTIONS)]["status"].unique()
            if len(invalid_status) > 0:
                st.warning(f"⚠️ Invalid status values found: {list(invalid_status)} — they will be set to 'Not Started'")
                upload_df.loc[~upload_df["status"].isin(STATUS_OPTIONS), "status"] = "Not Started"

            # Fill missing columns
            for col in ["description", "assignee", "milestone", "notes"]:
                if col not in upload_df.columns:
                    upload_df[col] = ""
            for col in ["budget_allocated", "budget_spent"]:
                if col not in upload_df.columns:
                    upload_df[col] = 0
                upload_df[col] = pd.to_numeric(upload_df[col], errors="coerce").fillna(0)

            st.markdown("### Preview")
            st.dataframe(upload_df[["deliverable","assignee","due_date","status","milestone"]], 
                        use_container_width=True, hide_index=True)

            if st.button("✅ Import to Tracker", type="primary", use_container_width=True):
                if "Replace" in import_mode:
                    db["deliverables"].delete_many({})
                    start_id = 1
                else:
                    last = db["deliverables"].find_one(sort=[("id", -1)])
                    start_id = (last["id"] + 1) if last else 1

                upload_df["id"] = range(start_id, start_id + len(upload_df))
                records = upload_df[[c for c in COLUMNS if c in upload_df.columns]].to_dict("records")
                db["deliverables"].insert_many(records)

                st.success(f"🎉 {len(records)} deliverables imported successfully!")
                st.balloons()

        except Exception as e:
            st.error(f"Could not read file: {e}")

# ═══════════════════════════════════════════════════════════════════════
# TAB 3 — EXPORT CURRENT
# ═══════════════════════════════════════════════════════════════════════
with tab3:
    st.subheader("📤 Export Current Deliverables")
    st.caption("Download your current deliverables as Excel, edit them, then re-upload to update.")

    if existing_df.empty:
        st.info("No deliverables yet.")
    else:
        st.dataframe(
            existing_df[[c for c in ["deliverable","assignee","status","due_date","milestone"] if c in existing_df.columns]].head(10),
            use_container_width=True, hide_index=True
        )
        st.markdown(f"**{len(existing_df)} total deliverables**")

        try:
            export_buffer = create_export_excel(existing_df)
            st.download_button(
                "⬇️ Download Current Deliverables (Excel)",
                data=export_buffer,
                file_name=f"deliverables_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )
            st.info("💡 Edit this file in Excel, then re-upload it in the **Upload & Import** tab using **Replace ALL** mode to update everything at once!")
        except ImportError:
            csv = existing_df.to_csv(index=False).encode("utf-8")
            st.download_button("⬇️ Download as CSV instead",
                data=csv, file_name=f"deliverables_{date.today()}.csv", mime="text/csv")

# ═══════════════════════════════════════════════════════════════════════
# GANTT CHART
# ═══════════════════════════════════════════════════════════════════════

st.markdown("---")
st.subheader("📊 Gantt Chart")
st.caption("Visual timeline of all your deliverables.")

if existing_df.empty:
    st.info("No deliverables yet — add some to see the Gantt chart!")
else:
    # Filter controls
    col1, col2 = st.columns(2)
    with col1:
        milestones = ["All"] + sorted(existing_df["milestone"].dropna().unique().tolist())
        gantt_milestone = st.selectbox("Filter by Milestone", milestones, key="gantt_ms")
    with col2:
        gantt_status = st.multiselect("Filter by Status", STATUS_OPTIONS, default=STATUS_OPTIONS, key="gantt_st")

    gantt_df = existing_df.copy()
    if gantt_milestone != "All":
        gantt_df = gantt_df[gantt_df["milestone"] == gantt_milestone]
    gantt_df = gantt_df[gantt_df["status"].isin(gantt_status)]
    gantt_df = gantt_df.dropna(subset=["due_date"]).sort_values("due_date")

    if gantt_df.empty:
        st.info("No deliverables match your filters.")
    else:
        # Build Gantt data for Chart.js
        status_colors = {
            "Not Started": "#888780",
            "In Progress": "#378ADD",
            "Under Review": "#BA7517",
            "Complete": "#639922",
            "Blocked": "#E24B4A",
        }

        gantt_data = []
        for _, row in gantt_df.iterrows():
            try:
                due = datetime.strptime(str(row["due_date"]), "%Y-%m-%d")
                gantt_data.append({
                    "label": str(row["deliverable"])[:40],
                    "due": row["due_date"],
                    "status": str(row["status"]),
                    "assignee": str(row.get("assignee", "")),
                    "milestone": str(row.get("milestone", "")),
                    "color": status_colors.get(str(row["status"]), "#888780")
                })
            except:
                pass

        gantt_json = json.dumps(gantt_data)
        today_str = str(date.today())

        gantt_html = f"""
<div id="gantt-wrap" style="width:100%; overflow-x:auto; padding: 0.5rem 0;">
<canvas id="ganttChart"></canvas>
</div>

<div style="display:flex; flex-wrap:wrap; gap:12px; margin-top:12px; font-size:12px; color:var(--color-text-secondary);">
  <span style="display:flex;align-items:center;gap:4px;"><span style="width:10px;height:10px;border-radius:2px;background:#888780;"></span>Not Started</span>
  <span style="display:flex;align-items:center;gap:4px;"><span style="width:10px;height:10px;border-radius:2px;background:#378ADD;"></span>In Progress</span>
  <span style="display:flex;align-items:center;gap:4px;"><span style="width:10px;height:10px;border-radius:2px;background:#BA7517;"></span>Under Review</span>
  <span style="display:flex;align-items:center;gap:4px;"><span style="width:10px;height:10px;border-radius:2px;background:#639922;"></span>Complete</span>
  <span style="display:flex;align-items:center;gap:4px;"><span style="width:10px;height:10px;border-radius:2px;background:#E24B4A;"></span>Blocked</span>
</div>

<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<script>
const data = {gantt_json};
const today = new Date('{today_str}');

const labels = data.map(d => d.label);
const colors = data.map(d => d.color);

const minDate = data.reduce((m, d) => d.due < m ? d.due : m, data[0].due);
const maxDate = data.reduce((m, d) => d.due > m ? d.due : m, data[0].due);

const startDate = new Date(minDate);
startDate.setMonth(startDate.getMonth() - 1);
const endDate = new Date(maxDate);
endDate.setMonth(endDate.getMonth() + 1);

const startMs = startDate.getTime();
const totalMs = endDate.getTime() - startMs;

const chartData = data.map(d => {{
  const due = new Date(d.due).getTime();
  const barEnd = due;
  const barStart = startMs;
  return [barStart, barEnd];
}});

const wrap = document.getElementById('gantt-wrap');
const height = Math.max(300, data.length * 44 + 80);
wrap.style.height = height + 'px';

const ctx = document.getElementById('ganttChart').getContext('2d');
const chart = new Chart(ctx, {{
  type: 'bar',
  data: {{
    labels: labels,
    datasets: [{{
      data: chartData,
      backgroundColor: colors,
      borderRadius: 4,
      borderSkipped: false,
    }}]
  }},
  options: {{
    indexAxis: 'y',
    responsive: true,
    maintainAspectRatio: false,
    plugins: {{
      legend: {{ display: false }},
      tooltip: {{
        callbacks: {{
          title: (items) => data[items[0].dataIndex].label,
          label: (item) => {{
            const d = data[item.dataIndex];
            return [
              'Due: ' + d.due,
              'Status: ' + d.status,
              d.assignee ? 'Assignee: ' + d.assignee : '',
              d.milestone ? 'Milestone: ' + d.milestone : ''
            ].filter(Boolean);
          }}
        }}
      }}
    }},
    scales: {{
      x: {{
        type: 'linear',
        min: startMs,
        max: endDate.getTime(),
        ticks: {{
          callback: (val) => {{
            const d = new Date(val);
            return d.toLocaleDateString('en-US', {{month: 'short', year: 'numeric'}});
          }},
          maxTicksLimit: 8,
          color: '#888'
        }},
        grid: {{ color: 'rgba(128,128,128,0.1)' }}
      }},
      y: {{
        ticks: {{
          font: {{ size: 11 }},
          color: '#666'
        }},
        grid: {{ display: false }}
      }}
    }}
  }},
  plugins: [{{
    id: 'todayLine',
    afterDraw(chart) {{
      const ctx = chart.ctx;
      const xAxis = chart.scales.x;
      const yAxis = chart.scales.y;
      const todayMs = today.getTime();
      if (todayMs >= xAxis.min && todayMs <= xAxis.max) {{
        const x = xAxis.getPixelForValue(todayMs);
        ctx.save();
        ctx.beginPath();
        ctx.moveTo(x, yAxis.top);
        ctx.lineTo(x, yAxis.bottom);
        ctx.lineWidth = 2;
        ctx.strokeStyle = '#E24B4A';
        ctx.setLineDash([4, 4]);
        ctx.stroke();
        ctx.fillStyle = '#E24B4A';
        ctx.font = '11px sans-serif';
        ctx.fillText('Today', x + 4, yAxis.top + 12);
        ctx.restore();
      }}
    }}
  }}]
}});
</script>
"""
        st.components.v1.html(gantt_html, height=len(gantt_data) * 44 + 180, scrolling=False)

        # Also export Gantt as table
        st.markdown("---")
        st.markdown("**📋 Gantt Table View**")
        gantt_table = gantt_df[["deliverable","milestone","assignee","status","due_date"]].copy()
        gantt_table.columns = ["Deliverable","Milestone","Assignee","Status","Due Date"]
        st.dataframe(gantt_table, use_container_width=True, hide_index=True)
